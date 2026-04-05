#!/usr/bin/env python3
"""
export_excel.py — reddit-intel skill Excel 导出脚本
Phase 4：生成双 Sheet 的 Excel 报告

Sheet 1「舆情数据」：18列，每条帖子一行，带颜色编码
Sheet 2「概览」：基础统计 + 情绪分布图 + 内容质量 + 分类分布
               + 高价值信号 + 竞品提及 + 综合行动建议

注意：Sheet 2 的所有数字由 Python 计算，不允许估算。
"""

import argparse
import json
import sys
import re
from collections import Counter
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
except ImportError:
    print("错误：缺少 openpyxl 库，请运行：pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ─────────────────────────────────────────────
# 颜色常量（与 OUTPUT_SCHEMA.md 保持一致）
# ─────────────────────────────────────────────
C = {
    # 表头
    "header_bg":   "1a1a2e",
    "header_fg":   "ffffff",
    # 情绪
    "sent_pos_bg": "d4edda",  "sent_pos_fg": "155724",
    "sent_neg_bg": "f8d7da",  "sent_neg_fg": "721c24",
    "sent_neu_bg": "e2e3e5",  "sent_neu_fg": "383d41",
    "sent_mix_bg": "fff3cd",  "sent_mix_fg": "856404",
    # 内容类型
    "ct_valid_bg": "ffffff",  "ct_valid_fg": "1a1a1a",
    "ct_other_bg": "e2e3e5",  "ct_other_fg": "6c757d",
    "ct_promo_bg": "fff3cd",  "ct_promo_fg": "856404",
    # 置信度
    "conf_h_bg":   "d4edda",  "conf_h_fg":   "155724",
    "conf_m_bg":   "fff3cd",  "conf_m_fg":   "856404",
    "conf_l_bg":   "f8d7da",  "conf_l_fg":   "721c24",
    # 行动点
    "action_bg":   "fefce8",
    # 奇偶行
    "row_odd":     "ffffff",
    "row_even":    "f8f9fa",
    # URL
    "url_fg":      "0563c1",
    # Sheet 2 区块标题
    "block_bg":    "1a1a2e",
    "block_fg":    "ffffff",
}

SENTIMENT_MAP = {
    "正面": ("sent_pos_bg", "sent_pos_fg"),
    "负面": ("sent_neg_bg", "sent_neg_fg"),
    "中性": ("sent_neu_bg", "sent_neu_fg"),
    "混合": ("sent_mix_bg", "sent_mix_fg"),
}

CONTENT_TYPE_MAP = {
    "有效文字": ("ct_valid_bg", "ct_valid_fg"),
    "图片帖":   ("ct_other_bg", "ct_other_fg"),
    "自动播报": ("ct_other_bg", "ct_other_fg"),
    "营销推广": ("ct_promo_bg", "ct_promo_fg"),
}

CONFIDENCE_MAP = {
    "高": ("conf_h_bg", "conf_h_fg"),
    "中": ("conf_m_bg", "conf_m_fg"),
    "低": ("conf_l_bg", "conf_l_fg"),
}


# ─────────────────────────────────────────────
# 样式工具函数
# ─────────────────────────────────────────────
def make_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)


def make_font(hex_color, bold=False, size=10, underline=None):
    kwargs = dict(color=hex_color, bold=bold, size=size)
    if underline:
        kwargs["underline"] = underline
    return Font(**kwargs)


def make_alignment(horizontal="left", wrap_text=True, vertical="top"):
    return Alignment(horizontal=horizontal, wrap_text=wrap_text, vertical=vertical)


def thin_border():
    side = Side(style="thin", color="d0d0d0")
    return Border(bottom=side)


def apply_header_row(ws, headers):
    """写入表头行，设置深色背景白色字体。"""
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.fill = make_fill(C["header_bg"])
        cell.font = make_font(C["header_fg"], bold=True)
        cell.alignment = make_alignment(horizontal="center", wrap_text=False)


def apply_block_header(ws, row, col, text, end_col=None):
    """Sheet 2 区块标题行。"""
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = make_fill(C["block_bg"])
    cell.font = make_font(C["block_fg"], bold=True, size=11)
    cell.alignment = make_alignment(horizontal="left", wrap_text=False)
    if end_col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=end_col)
    return row + 1


# ─────────────────────────────────────────────
# Sheet 1：舆情数据
# ─────────────────────────────────────────────
SHEET1_COLUMNS = [
    # (列名, JSON字段, 宽度, 对齐, 换行)
    ("标题（原文）",       "title_original",          50, "left",   True),
    ("标题（中文）",       "title_zh",                50, "left",   True),
    ("内容摘要（中文）",   "summary_zh",              70, "left",   True),
    ("内容类型",           "content_type",            14, "center", False),
    ("核心意图",           "core_intent",             14, "center", False),
    ("主分类",             "main_category",           18, "center", False),
    ("子标签",             "sub_tags",                24, "left",   False),
    ("情绪",               "sentiment",               10, "center", False),
    ("提及竞品/项目",      "competitors_mentioned",   24, "left",   False),
    ("使用场景",           "use_case",                28, "left",   True),
    ("热评精华",           "top_comments_summary",    60, "left",   True),
    ("行动点",             "action_item",             50, "left",   True),
    ("置信度",             "confidence",              10, "center", False),
    ("点赞数",             "upvotes",                  8, "center", False),
    ("评论数",             "num_comments",             8, "center", False),
    ("来源版块",           "subreddit",               18, "center", False),
    ("发帖时间",           "created_date",            12, "center", False),
    ("原帖链接",           "url",                     40, "left",   False),
]


def write_sheet1(wb, posts):
    """生成 Sheet 1：舆情数据。"""
    ws = wb.active
    ws.title = "舆情数据"

    # 冻结第一行
    ws.freeze_panes = "A2"

    # 表头
    headers = [col[0] for col in SHEET1_COLUMNS]
    apply_header_row(ws, headers)

    # 设置列宽
    for col_idx, (_, _, width, _, _) in enumerate(SHEET1_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 数据行
    for row_idx, post in enumerate(posts, 2):
        bg_color = C["row_odd"] if row_idx % 2 == 1 else C["row_even"]

        for col_idx, (_, field, _, align, wrap) in enumerate(SHEET1_COLUMNS, 1):
            # 取值
            raw = post.get(field, "")
            if isinstance(raw, list):
                value = "、".join(str(v) for v in raw) if raw else ""
            elif raw is None:
                value = ""
            else:
                value = raw

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # 基础样式
            cell.fill = make_fill(bg_color)
            cell.alignment = make_alignment(horizontal=align, wrap_text=wrap)
            cell.font = make_font("1a1a1a", size=10)

            # 特殊列样式覆盖
            if field == "sentiment" and value in SENTIMENT_MAP:
                bg_key, fg_key = SENTIMENT_MAP[value]
                cell.fill = make_fill(C[bg_key])
                cell.font = make_font(C[fg_key], size=10)

            elif field == "content_type" and value in CONTENT_TYPE_MAP:
                bg_key, fg_key = CONTENT_TYPE_MAP[value]
                cell.fill = make_fill(C[bg_key])
                cell.font = make_font(C[fg_key], size=10)

            elif field == "confidence" and value in CONFIDENCE_MAP:
                bg_key, fg_key = CONFIDENCE_MAP[value]
                cell.fill = make_fill(C[bg_key])
                cell.font = make_font(C[fg_key], size=10)

            elif field == "action_item" and value:
                cell.fill = make_fill(C["action_bg"])
                cell.font = make_font("1a1a1a", bold=True, size=10)

            elif field == "url" and value:
                cell.font = make_font(C["url_fg"], size=10, underline="single")
                cell.hyperlink = value

            elif field in ("upvotes", "num_comments"):
                cell.alignment = make_alignment(horizontal="center", wrap_text=False)
                if isinstance(value, (int, float)):
                    cell.number_format = "#,##0"

        # 最小行高
        ws.row_dimensions[row_idx].height = 30

    return ws


# ─────────────────────────────────────────────
# Sheet 2：概览（Python 计算，不估算）
# ─────────────────────────────────────────────
def write_sheet2(wb, posts, actions_text, keyword):
    """生成 Sheet 2：概览。所有统计数字由 Python 计算。"""
    ws = wb.create_sheet(title="概览")

    # 列宽
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 50

    row = 1

    # ── 区块1：基础统计 ──
    row = apply_block_header(ws, row, 1, "📊 基础统计", end_col=5)

    total = len(posts)
    valid = sum(1 for p in posts if p.get("content_type") == "有效文字")
    dates = [p.get("created_date", "") for p in posts if p.get("created_date")]
    subreddits = [p.get("subreddit", "") for p in posts if p.get("subreddit")]
    unique_subs = sorted(set(subreddits))

    stats = [
        ("总帖数", total),
        ("有效帖数（有效文字）", valid),
        ("时间跨度", f"{min(dates)} 至 {max(dates)}" if dates else "未知"),
        ("来源版块数", len(unique_subs)),
    ]
    for label, value in stats:
        ws.cell(row=row, column=1, value=label).font = make_font("1a1a1a", bold=True)
        cell_v = ws.cell(row=row, column=2, value=value)
        cell_v.alignment = make_alignment(horizontal="right", wrap_text=False)
        row += 1

    # 版块分布 Top 10
    sub_counter = Counter(subreddits).most_common(10)
    if sub_counter:
        row += 1
        ws.cell(row=row, column=1, value="版块分布（Top 10）").font = make_font("1a1a1a", bold=True)
        row += 1
        for sub, cnt in sub_counter:
            ws.cell(row=row, column=1, value=sub).font = make_font("555555")
            cell_cnt = ws.cell(row=row, column=2, value=cnt)
            cell_cnt.alignment = make_alignment(horizontal="right", wrap_text=False)
            row += 1

    row += 1

    # ── 区块2：情绪分布 ──
    row = apply_block_header(ws, row, 1, "😊 情绪分布", end_col=5)

    sentiments = [p.get("sentiment", "") for p in posts if p.get("sentiment")]
    sent_counter = Counter(sentiments)
    sent_order = ["正面", "负面", "中性", "混合"]
    sent_data_row_start = row

    for s in sent_order:
        cnt = sent_counter.get(s, 0)
        pct = cnt / total * 100 if total else 0
        ws.cell(row=row, column=1, value=s).font = make_font("1a1a1a")
        cell_cnt = ws.cell(row=row, column=2, value=cnt)
        cell_cnt.alignment = make_alignment(horizontal="right", wrap_text=False)
        cell_pct = ws.cell(row=row, column=3, value=round(pct, 1))
        cell_pct.number_format = '0.0"%"'
        cell_pct.alignment = make_alignment(horizontal="right", wrap_text=False)
        row += 1

    # 简单条形图
    chart_row = sent_data_row_start
    chart = BarChart()
    chart.type = "bar"
    chart.title = "情绪分布"
    chart.y_axis.title = "情绪类型"
    chart.x_axis.title = "帖子数"
    chart.width = 12
    chart.height = 8

    data_ref = Reference(ws,
                         min_col=2, min_row=sent_data_row_start,
                         max_col=2, max_row=sent_data_row_start + len(sent_order) - 1)
    cats_ref = Reference(ws,
                         min_col=1, min_row=sent_data_row_start,
                         max_row=sent_data_row_start + len(sent_order) - 1)
    chart.add_data(data_ref)
    chart.set_categories(cats_ref)
    chart.series[0].title = None

    ws.add_chart(chart, f"D{chart_row}")
    row += 1

    # ── 区块3：内容质量 ──
    row = apply_block_header(ws, row, 1, "📋 内容质量", end_col=5)

    ct_counter = Counter(p.get("content_type", "") for p in posts)
    for ct in ["有效文字", "图片帖", "自动播报", "营销推广"]:
        cnt = ct_counter.get(ct, 0)
        pct = cnt / total * 100 if total else 0
        ws.cell(row=row, column=1, value=ct).font = make_font("1a1a1a")
        cell_cnt = ws.cell(row=row, column=2, value=cnt)
        cell_cnt.alignment = make_alignment(horizontal="right", wrap_text=False)
        cell_pct = ws.cell(row=row, column=3, value=round(pct, 1))
        cell_pct.number_format = '0.0"%"'
        cell_pct.alignment = make_alignment(horizontal="right", wrap_text=False)
        row += 1

    # 低置信度统计
    low_conf = sum(1 for p in posts if p.get("confidence") == "低")
    low_pct = low_conf / total * 100 if total else 0
    row += 1
    ws.cell(row=row, column=1, value="低置信度帖子数").font = make_font("721c24", bold=True)
    cell_lc = ws.cell(row=row, column=2, value=low_conf)
    cell_lc.alignment = make_alignment(horizontal="right", wrap_text=False)
    cell_lcp = ws.cell(row=row, column=3, value=round(low_pct, 1))
    cell_lcp.number_format = '0.0"%"'
    cell_lcp.alignment = make_alignment(horizontal="right", wrap_text=False)
    row += 2

    # ── 区块4：分类分布 ──
    row = apply_block_header(ws, row, 1, "🏷️ 分类分布", end_col=5)

    # 主分类 Top 10
    main_cats = [p.get("main_category", "") for p in posts if p.get("main_category")]
    cat_counter = Counter(main_cats).most_common(10)
    ws.cell(row=row, column=1, value="主分类（Top 10）").font = make_font("1a1a1a", bold=True)
    row += 1
    for cat, cnt in cat_counter:
        ws.cell(row=row, column=1, value=cat).font = make_font("555555")
        cell_cnt = ws.cell(row=row, column=2, value=cnt)
        cell_cnt.alignment = make_alignment(horizontal="right", wrap_text=False)
        row += 1

    # 核心意图分布
    row += 1
    intents = [p.get("core_intent", "") for p in posts if p.get("core_intent")]
    intent_counter = Counter(intents).most_common()
    ws.cell(row=row, column=1, value="核心意图分布").font = make_font("1a1a1a", bold=True)
    row += 1
    for intent, cnt in intent_counter:
        ws.cell(row=row, column=1, value=intent).font = make_font("555555")
        cell_cnt = ws.cell(row=row, column=2, value=cnt)
        cell_cnt.alignment = make_alignment(horizontal="right", wrap_text=False)
        row += 1

    row += 1

    # ── 区块5：高价值信号 ──
    row = apply_block_header(ws, row, 1, "🔥 高价值信号", end_col=5)

    # 点赞数最高 Top 5
    ws.cell(row=row, column=1, value="点赞数 Top 5").font = make_font("1a1a1a", bold=True)
    row += 1
    top_upvoted = sorted(posts, key=lambda x: x.get("upvotes", 0), reverse=True)[:5]
    for p in top_upvoted:
        title = p.get("title_zh") or p.get("title_original", "")
        title = (title[:60] + "...") if len(title) > 60 else title
        upvotes = p.get("upvotes", 0)
        sentiment = p.get("sentiment", "")
        action = p.get("action_item", "")

        ws.cell(row=row, column=1, value=title).font = make_font("1a1a1a")
        cell_up = ws.cell(row=row, column=2, value=upvotes)
        cell_up.alignment = make_alignment(horizontal="right", wrap_text=False)
        ws.cell(row=row, column=3, value=sentiment).font = make_font("555555")
        action_cell = ws.cell(row=row, column=4, value=action)
        action_cell.alignment = make_alignment(wrap_text=True)
        action_cell.font = make_font("1a1a1a")
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        row += 1

    # 评论数最多 Top 5
    row += 1
    ws.cell(row=row, column=1, value="评论数 Top 5").font = make_font("1a1a1a", bold=True)
    row += 1
    top_comments = sorted(posts, key=lambda x: x.get("num_comments", 0), reverse=True)[:5]
    for p in top_comments:
        title = p.get("title_zh") or p.get("title_original", "")
        title = (title[:60] + "...") if len(title) > 60 else title
        num_c = p.get("num_comments", 0)

        ws.cell(row=row, column=1, value=title).font = make_font("1a1a1a")
        cell_nc = ws.cell(row=row, column=2, value=num_c)
        cell_nc.alignment = make_alignment(horizontal="right", wrap_text=False)
        row += 1

    # 情绪最强负面 Top 3
    row += 1
    ws.cell(row=row, column=1, value="负面情绪 Top 3").font = make_font("721c24", bold=True)
    row += 1
    neg_posts = [p for p in posts if p.get("sentiment") == "负面"]
    neg_posts_sorted = sorted(neg_posts, key=lambda x: x.get("upvotes", 0), reverse=True)[:3]
    for p in neg_posts_sorted:
        title = p.get("title_zh") or p.get("title_original", "")
        title = (title[:60] + "...") if len(title) > 60 else title
        summary = p.get("summary_zh", "")[:100]

        ws.cell(row=row, column=1, value=title).font = make_font("721c24")
        summary_cell = ws.cell(row=row, column=2, value=summary)
        summary_cell.alignment = make_alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 1

    row += 1

    # ── 区块6：竞品提及（仅在有数据时显示）──
    all_competitors = []
    for p in posts:
        comps = p.get("competitors_mentioned", [])
        if isinstance(comps, list):
            all_competitors.extend(c for c in comps if c)
        elif isinstance(comps, str) and comps:
            all_competitors.extend(c.strip() for c in comps.split("、") if c.strip())

    if all_competitors:
        row = apply_block_header(ws, row, 1, "🏢 竞品/项目提及（Top 10）", end_col=5)
        comp_counter = Counter(all_competitors).most_common(10)
        for comp, cnt in comp_counter:
            ws.cell(row=row, column=1, value=comp).font = make_font("1a1a1a")
            cell_cnt = ws.cell(row=row, column=2, value=cnt)
            cell_cnt.alignment = make_alignment(horizontal="right", wrap_text=False)
            row += 1
        row += 1

    # ── 区块7：综合行动建议（Claude 生成文字传入）──
    row = apply_block_header(ws, row, 1, "💡 综合行动建议", end_col=5)

    if actions_text:
        # 按换行分割，每条作为单独行
        action_lines = [line.strip() for line in actions_text.split("\n") if line.strip()]
        for i, line in enumerate(action_lines, 1):
            cell = ws.cell(row=row, column=1, value=f"{i}. {line}")
            cell.font = make_font("1a1a1a", size=10)
            cell.alignment = make_alignment(wrap_text=True)
            cell.fill = make_fill(C["action_bg"])
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws.row_dimensions[row].height = 45
            row += 1
    else:
        ws.cell(row=row, column=1, value="（未提供行动建议）").font = make_font("999999")
        row += 1

    return ws


# ─────────────────────────────────────────────
# 主逻辑
# ─────────────────────────────────────────────
def sanitize_keyword(keyword):
    """清理关键词用于文件名。"""
    cleaned = re.sub(r'[^\w\u4e00-\u9fff]', '', keyword)
    return cleaned[:20] if cleaned else "keyword"


def main():
    parser = argparse.ArgumentParser(
        description="Export reddit-intel processed posts to Excel",
    )
    parser.add_argument("--input", default="posts_processed.json",
                        help="已处理的帖子 JSON 文件路径")
    parser.add_argument("--keyword", default="reddit",
                        help="用于文件命名的关键词")
    parser.add_argument("--actions", default="",
                        help="综合行动建议文字（换行分隔的多条建议）")
    parser.add_argument("--output", default="",
                        help="输出文件路径（默认按规则自动生成）")

    args = parser.parse_args()

    # 读取数据
    print(f"[读取] {args.input}", file=sys.stderr)
    try:
        with open(args.input, "r", encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"错误：找不到文件 {args.input}", file=sys.stderr)
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"错误：JSON 解析失败: {e}", file=sys.stderr)
        sys.exit(1)

    # 兼容两种数据格式：纯列表 或 {posts: [...]}
    if isinstance(data, list):
        posts = data
    elif isinstance(data, dict):
        posts = data.get("posts", data.get("data", []))
    else:
        print("错误：无法识别的数据格式", file=sys.stderr)
        sys.exit(1)

    if not posts:
        print("警告：帖子列表为空，生成空报告", file=sys.stderr)

    print(f"[统计] 共 {len(posts)} 条帖子", file=sys.stderr)

    # 确定输出文件名
    if args.output:
        output_path = args.output
    else:
        today = datetime.now().strftime("%Y%m%d")
        kw = sanitize_keyword(args.keyword)
        output_path = f"reddit_intel_{kw}_{today}.xlsx"

    # 创建工作簿
    wb = Workbook()

    # Sheet 1
    print("[生成] Sheet 1：舆情数据...", file=sys.stderr)
    write_sheet1(wb, posts)

    # Sheet 2
    print("[生成] Sheet 2：概览...", file=sys.stderr)
    write_sheet2(wb, posts, args.actions, args.keyword)

    # 保存
    wb.save(output_path)

    # 统计输出
    total = len(posts)
    conf_h = sum(1 for p in posts if p.get("confidence") == "高")
    conf_m = sum(1 for p in posts if p.get("confidence") == "中")
    conf_l = sum(1 for p in posts if p.get("confidence") == "低")

    print(f"\n✅ Excel 已生成: {output_path}", file=sys.stderr)
    print(f"   Sheet 1「舆情数据」: {total} 条帖子，18 列", file=sys.stderr)
    print(f"   Sheet 2「概览」: 基础统计 / 情绪分布 / 内容质量 / 分类分布 / 高价值信号 / 综合建议", file=sys.stderr)
    print(f"   置信度分布: 高 {conf_h} 条 / 中 {conf_m} 条 / 低 {conf_l} 条", file=sys.stderr)

    # 标准输出供 Claude 读取
    print(json.dumps({
        "status": "ok",
        "output_file": output_path,
        "total_posts": total,
        "confidence": {"高": conf_h, "中": conf_m, "低": conf_l},
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
